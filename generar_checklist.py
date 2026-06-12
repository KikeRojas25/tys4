"""
Genera CHECKLIST_COMERCIAL.xlsx y CHECKLIST_COMERCIAL.pdf
a partir de la estructura del checklist del módulo Comercial.
"""
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.dimensions import RowDimension

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm, mm
from reportlab.platypus import (
    SimpleDocTemplate,
    Paragraph,
    Spacer,
    Table,
    TableStyle,
    PageBreak,
)


# ────────────────────────────────────────────────────────────────────
# Estructura del checklist
# ────────────────────────────────────────────────────────────────────

SECCIONES = [
    {
        "titulo": "1. Lead Times Comerciales (/comercial/leadtimes)",
        "descripcion": "Mantenimiento de tiempos de entrega (días) por cliente y provincia destino.",
        "casos": [
            ("1.1", "Ingresar a la pantalla", "Carga el dropdown 'Cliente' con la lista de clientes asignados al usuario"),
            ("1.2", "Seleccionar un cliente", "Se cargan todas las provincias destino agrupadas por departamento"),
            ("1.3", "Verificar agrupación visual", "Cada departamento aparece una sola vez (rowspan) y debajo lista sus provincias"),
            ("1.4", "Filtrar por texto (departamento o provincia)", "El listado se reduce dinámicamente y resalta la coincidencia"),
            ("1.5", "Activar 'Solo vacíos'", "Solo se muestran las provincias sin días asignados"),
            ("1.6", "Hacer clic sobre un número de día (1-30)", "El día queda marcado como seleccionado para esa provincia"),
            ("1.7", "Hacer clic sobre el día ya seleccionado", "El día se desmarca (vuelve a vacío)"),
            ("1.8", "Pulsar 'Guardar'", "Toast verde 'Lead times guardados correctamente (N registros)'"),
            ("1.9", "Recargar la pantalla y volver a seleccionar el mismo cliente", "Los días guardados persisten"),
            ("1.10", "Cambiar de cliente sin guardar", "Los cambios no se mezclan entre clientes"),
        ],
    },
    {
        "titulo": "2. Lead Times Operativos (/comercial/leadtimes/operativo)",
        "descripcion": "Tiempos de entrega estándar por provincia (no depende de cliente).",
        "casos": [
            ("2.1", "Ingresar a la pantalla", "Carga directa de todas las provincias (sin selector de cliente)"),
            ("2.2", "Filtrar y 'solo vacíos'", "Mismo comportamiento que el comercial"),
            ("2.3", "Seleccionar/deseleccionar día", "Funciona igual que el flujo comercial"),
            ("2.4", "Pulsar 'Guardar'", "Toast verde con el total de registros afectados"),
            ("2.5", "Verificar persistencia", "Al recargar, los valores guardados se mantienen"),
        ],
    },
    {
        "titulo": "3.0 Registro de Citas — Selección base (/comercial/registro-citas)",
        "descripcion": "Pantalla maestra con 4 flujos: Registro de OR, Cita de OT, Reclamos, Instrucción de Incidencias.",
        "casos": [
            ("3.0.1", "Ingresar a la pantalla", "Se muestra el selector de cliente + tarjetas de tipo (Registro de OR, Cita de OT, Reclamos, Instrucción de Incidencias)"),
            ("3.0.2", "Seleccionar cliente", "Se cargan internamente las OTs pendientes del cliente"),
            ("3.0.3", "Cambiar de cliente", "Se resetean los campos de detalle y selecciones previas"),
        ],
    },
    {
        "titulo": "3.1 Registro de OR (Orden de Recojo)",
        "descripcion": "Registro de una nueva orden de recojo desde Registro de Citas.",
        "casos": [
            ("3.1.1", "Seleccionar tipo 'Registro de OR'", "Se muestran campos: fecha y hora de cita, origen, contacto, teléfono, destino, centro acopio, observaciones, destinos finales"),
            ("3.1.2", "Fecha de recojo por defecto", "Día siguiente a hoy"),
            ("3.1.3", "Botones para ajustar la fecha (+/-)", "No permite seleccionar fechas anteriores a hoy (toast de aviso)"),
            ("3.1.4", "Hora con formato inválido", "Toast 'Hora inválida (HH:mm)'"),
            ("3.1.5", "Agregar destino final", "Requiere destino y cantidad; se agrega a la tabla con peso/volumen opcionales"),
            ("3.1.6", "Eliminar destino final", "Se quita la fila de la tabla"),
            ("3.1.7", "Totales (cantidad / peso / volumen)", "Se actualizan en pie de tabla"),
            ("3.1.8", "Guardar con campos obligatorios faltantes", "Toast amarillo 'Complete los campos obligatorios del recojo (*)'"),
            ("3.1.9", "Guardar correctamente", "Modal de confirmación → toast 'Orden de recojo registrada correctamente'"),
            ("3.1.10", "Toggle 'Listado de OR embebido'", "Se despliega el listado de órdenes de recojo del cliente"),
        ],
    },
    {
        "titulo": "3.2 Cita de OT",
        "descripcion": "Agendar cita para una OT pendiente del cliente.",
        "casos": [
            ("3.2.1", "Seleccionar tipo 'Cita de OT'", "Se muestra la tabla de OTs pendientes del cliente"),
            ("3.2.2", "Buscar por numcp/remitente/destinatario/destino", "Filtro client-side dinámico"),
            ("3.2.3", "Seleccionar una OT de la tabla", "Queda marcada como selectedOT"),
            ("3.2.4", "Guardar sin OT seleccionada", "Toast 'Debe seleccionar una OT de la tabla'"),
            ("3.2.5", "Guardar sin fecha o hora", "Toast 'Debe seleccionar fecha y hora de la cita'"),
            ("3.2.6", "Guardar correctamente", "Confirmación con resumen (numcp + fecha + hora) → toast 'Cita guardada correctamente'"),
        ],
    },
    {
        "titulo": "3.3 Registro de Reclamos",
        "descripcion": "Flujo de 4 áreas: Programación local, Programación provincia, Tráfico provincia, Almacén.",
        "casos": [
            ("3.3.1", "Seleccionar tipo 'Registro de Reclamos'", "Se muestran las 4 tarjetas de área"),
            ("3.3.2", "Programación local → seleccionar subtipo (no recojo, no LI, arribo tarde, falta drive, no procedimiento)", "Aparece buscador de OR (mín. 3 caracteres)"),
            ("3.3.3", "Buscar OR con menos de 3 caracteres", "Toast 'Ingrese al menos 3 caracteres'"),
            ("3.3.4", "Buscar OR válida y seleccionarla", "Queda asociada al reclamo"),
            ("3.3.5", "Programación provincia → seleccionar 'Otros' sin texto", "Toast 'Especifique el detalle en Otros'"),
            ("3.3.6", "Tráfico → seleccionar subtipo y buscar OT", "Buscador de OT por numcp"),
            ("3.3.7", "Almacén → seleccionar subtipo", "Algunos requieren búsqueda de OT (falta-ingreso, faltante-carga); otros no"),
            ("3.3.8", "Guardar sin área seleccionada", "Toast 'Seleccione el área'"),
            ("3.3.9", "Guardar sin OR/OT relacionada cuando aplica", "Toast pidiendo seleccionar OR / OT"),
            ("3.3.10", "Guardar correctamente", "Toast verde 'Reclamo registrado (#N)'"),
            ("3.3.11", "Toggle 'Reclamos embebidos'", "Se muestra el seguimiento de reclamos del cliente dentro de la pantalla"),
        ],
    },
    {
        "titulo": "3.4 Instrucción de Incidencias (sobre OT observada)",
        "descripcion": "Registro de instrucción ante OT con observaciones de entrega.",
        "casos": [
            ("3.4.1", "Seleccionar tipo 'Instrucción de Incidencias'", "Carga tabla de OTs observadas del cliente"),
            ("3.4.2", "Filtrar OTs observadas (numcp/remitente/destinatario/destino/tipoentrega/observación)", "Filtro client-side dinámico"),
            ("3.4.3", "Seleccionar una OT", "Queda marcada como selectedOTObservada"),
            ("3.4.4", "Guardar sin OT observada", "Toast 'Seleccione una OT con observaciones'"),
            ("3.4.5", "Guardar sin instrucción (observaciones vacías)", "Toast 'Escriba la instrucción de la observación'"),
            ("3.4.6", "Guardar correctamente", "Toast 'Instrucción registrada (#N)'"),
        ],
    },
    {
        "titulo": "4. Seguimiento de Reclamos (/comercial/seguimiento-reclamos)",
        "descripcion": "Listado, detalle y cambio de estado de reclamos.",
        "casos": [
            ("4.1", "Ingresar a la pantalla", "Cargan filtros (cliente, área, estado, fechas), default fechas: últimos 30 días"),
            ("4.2", "Acceder con ?idcliente=N en URL", "Cliente queda preseleccionado"),
            ("4.3", "Pulsar 'Buscar'", "Lista de reclamos según filtros"),
            ("4.4", "Cambiar área / estado / rango de fechas y buscar", "Listado se actualiza correctamente"),
            ("4.5", "'Limpiar filtros'", "Vuelve a default (últimos 30 días, sin cliente/área/estado)"),
            ("4.6", "Hacer clic en 'Ver detalle' de un reclamo", "Se abre diálogo con datos + historial completo"),
            ("4.7", "Pulsar 'Cambiar estado'", "Aparece diálogo con estados disponibles (excluye el actual) y campo comentario"),
            ("4.8", "Confirmar cambio de estado sin elegir nuevo estado", "Toast 'Seleccione el nuevo estado'"),
            ("4.9", "Cambiar estado correctamente", "Toast verde, se refresca detalle e historial, listado actualizado"),
            ("4.10", "Verificar colores de severidad de estados", "registrado=info, en-atencion=warning, resuelto=success, descartado=danger"),
        ],
    },
    {
        "titulo": "5. Integración con otros módulos",
        "descripcion": "Componentes embebidos y dependencias con módulos externos.",
        "casos": [
            ("5.1", "Desde Registro de Citas → toggle Listado de OR", "El componente embebido listarOrdenRecojoComponent carga las OR del cliente"),
            ("5.2", "Desde Registro de Citas → toggle Reclamos", "El componente embebido SeguimientoReclamosComponent carga reclamos del cliente"),
            ("5.3", "Verificar que clientes filtrados por usuario", "Solo aparecen los clientes asignados al usuario logueado"),
        ],
    },
    {
        "titulo": "6. Validaciones cruzadas y errores",
        "descripcion": "Pruebas de errores, sesión, permisos y entradas inválidas.",
        "casos": [
            ("6.1", "Pérdida de conexión durante guardar", "Toast rojo de error, sin perder los datos del formulario"),
            ("6.2", "Sesión vencida", "Redirige a login"),
            ("6.3", "Permisos insuficientes (si aplica)", "Mensaje claro o pantalla bloqueada"),
            ("6.4", "Caracteres especiales en observaciones", "Se guardan tal cual"),
            ("6.5", "Campos numéricos con valores no válidos (negativos, texto)", "Validación visible o rechazo"),
        ],
    },
]

OUT_DIR = Path(__file__).parent
XLSX_PATH = OUT_DIR / "CHECKLIST_COMERCIAL.xlsx"
PDF_PATH = OUT_DIR / "CHECKLIST_COMERCIAL.pdf"


# ────────────────────────────────────────────────────────────────────
# Excel
# ────────────────────────────────────────────────────────────────────

def generar_excel():
    wb = Workbook()

    # Estilos
    title_font = Font(name="Calibri", size=16, bold=True, color="FFFFFF")
    title_fill = PatternFill("solid", fgColor="1F4E78")
    section_font = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
    section_fill = PatternFill("solid", fgColor="2E75B6")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="305496")
    desc_font = Font(name="Calibri", size=10, italic=True, color="555555")
    cell_font = Font(name="Calibri", size=10)
    summary_font = Font(name="Calibri", size=11, bold=True)
    summary_fill = PatternFill("solid", fgColor="D9E1F2")

    thin = Side(border_style="thin", color="999999")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    align_wrap = Alignment(horizontal="left", vertical="top", wrap_text=True)
    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # ── Hoja "Datos generales" ──────────────────────────────────────
    ws_info = wb.active
    ws_info.title = "Datos"
    ws_info.column_dimensions["A"].width = 32
    ws_info.column_dimensions["B"].width = 60

    ws_info["A1"] = "Checklist de Conformidad — Módulo Comercial"
    ws_info["A1"].font = title_font
    ws_info["A1"].fill = title_fill
    ws_info.merge_cells("A1:B1")
    ws_info.row_dimensions[1].height = 30
    ws_info["A1"].alignment = Alignment(horizontal="center", vertical="center")

    info_rows = [
        ("Nombre del responsable", ""),
        ("Rol", ""),
        ("Fecha de prueba", ""),
        ("Ambiente (Dev / QA / Prod)", ""),
        ("", ""),
        ("Leyenda de estado", ""),
        ("Pendiente", "[ ]"),
        ("Conforme", "[x]"),
        ("Con observación", "[!]"),
        ("Falla bloqueante", "[F]"),
    ]
    for i, (k, v) in enumerate(info_rows, start=3):
        ws_info.cell(row=i, column=1, value=k).font = summary_font
        ws_info.cell(row=i, column=1).alignment = align_wrap
        ws_info.cell(row=i, column=2, value=v).alignment = align_wrap
        if k:
            ws_info.cell(row=i, column=1).border = border
            ws_info.cell(row=i, column=2).border = border
        if k == "Leyenda de estado":
            ws_info.cell(row=i, column=1).fill = summary_fill

    # ── Hoja "Casos de prueba" ──────────────────────────────────────
    ws = wb.create_sheet("Casos de prueba")
    headers = ["#", "Caso", "Resultado esperado", "Estado", "Observación"]
    col_widths = [8, 50, 60, 16, 50]

    for idx, w in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = w

    # Título
    ws["A1"] = "Casos de prueba — Módulo Comercial"
    ws["A1"].font = title_font
    ws["A1"].fill = title_fill
    ws.merge_cells("A1:E1")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    row = 3
    for sec in SECCIONES:
        # Banda de sección
        ws.cell(row=row, column=1, value=sec["titulo"]).font = section_font
        ws.cell(row=row, column=1).fill = section_fill
        ws.cell(row=row, column=1).alignment = Alignment(horizontal="left", vertical="center")
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
        ws.row_dimensions[row].height = 22
        row += 1

        # Descripción
        ws.cell(row=row, column=1, value=sec["descripcion"]).font = desc_font
        ws.cell(row=row, column=1).alignment = align_wrap
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
        ws.row_dimensions[row].height = 18
        row += 1

        # Encabezados
        for c, h in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=c, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = align_center
            cell.border = border
        ws.row_dimensions[row].height = 20
        row += 1

        # Casos
        for num, caso, esperado in sec["casos"]:
            ws.cell(row=row, column=1, value=num)
            ws.cell(row=row, column=2, value=caso)
            ws.cell(row=row, column=3, value=esperado)
            ws.cell(row=row, column=4, value="[ ]")
            ws.cell(row=row, column=5, value="")
            for c in range(1, 6):
                cell = ws.cell(row=row, column=c)
                cell.font = cell_font
                cell.alignment = align_wrap
                cell.border = border
            ws.row_dimensions[row].height = 32
            row += 1

        # Espaciador
        row += 1

    # Freeze header row
    ws.freeze_panes = "A2"

    # ── Hoja "Resumen" ──────────────────────────────────────────────
    ws_sum = wb.create_sheet("Resumen")
    ws_sum.column_dimensions["A"].width = 50
    ws_sum.column_dimensions["B"].width = 14
    ws_sum.column_dimensions["C"].width = 14
    ws_sum.column_dimensions["D"].width = 18
    ws_sum.column_dimensions["E"].width = 20

    ws_sum["A1"] = "Resumen de conformidad"
    ws_sum["A1"].font = title_font
    ws_sum["A1"].fill = title_fill
    ws_sum.merge_cells("A1:E1")
    ws_sum["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws_sum.row_dimensions[1].height = 28

    sum_headers = ["Sección", "Total casos", "Conformes", "Con observación", "Falla bloqueante"]
    for c, h in enumerate(sum_headers, start=1):
        cell = ws_sum.cell(row=3, column=c, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = align_center
        cell.border = border

    total_casos = 0
    r = 4
    for sec in SECCIONES:
        n = len(sec["casos"])
        total_casos += n
        ws_sum.cell(row=r, column=1, value=sec["titulo"]).alignment = align_wrap
        ws_sum.cell(row=r, column=2, value=n).alignment = align_center
        for c in range(3, 6):
            ws_sum.cell(row=r, column=c, value="").alignment = align_center
        for c in range(1, 6):
            ws_sum.cell(row=r, column=c).border = border
            ws_sum.cell(row=r, column=c).font = cell_font
        ws_sum.row_dimensions[r].height = 32
        r += 1

    # Total
    ws_sum.cell(row=r, column=1, value="TOTAL").font = summary_font
    ws_sum.cell(row=r, column=1).fill = summary_fill
    ws_sum.cell(row=r, column=2, value=total_casos).font = summary_font
    ws_sum.cell(row=r, column=2).fill = summary_fill
    ws_sum.cell(row=r, column=2).alignment = align_center
    for c in range(3, 6):
        ws_sum.cell(row=r, column=c, value="").fill = summary_fill
    for c in range(1, 6):
        ws_sum.cell(row=r, column=c).border = border

    # Bloque de aprobación
    r += 3
    ws_sum.cell(row=r, column=1, value="Conformidad final").font = section_font
    ws_sum.cell(row=r, column=1).fill = section_fill
    ws_sum.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
    ws_sum.cell(row=r, column=1).alignment = Alignment(horizontal="left", vertical="center")
    ws_sum.row_dimensions[r].height = 22
    r += 1
    aprob = [
        "[ ] Aprobado — todas las funcionalidades operan según lo esperado.",
        "[ ] Aprobado con observaciones — listadas en la sección anterior, no bloquean producción.",
        "[ ] Rechazado — existen fallas bloqueantes que impiden el pase a producción.",
    ]
    for txt in aprob:
        ws_sum.cell(row=r, column=1, value=txt).alignment = align_wrap
        ws_sum.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
        ws_sum.row_dimensions[r].height = 22
        r += 1

    r += 1
    ws_sum.cell(row=r, column=1, value="Firma del responsable:").font = summary_font
    r += 1
    ws_sum.cell(row=r, column=1, value="Fecha:").font = summary_font

    wb.save(XLSX_PATH)
    print(f"Excel generado: {XLSX_PATH}")


# ────────────────────────────────────────────────────────────────────
# PDF
# ────────────────────────────────────────────────────────────────────

def generar_pdf():
    doc = SimpleDocTemplate(
        str(PDF_PATH),
        pagesize=landscape(A4),
        leftMargin=12 * mm,
        rightMargin=12 * mm,
        topMargin=12 * mm,
        bottomMargin=12 * mm,
        title="Checklist Conformidad - Módulo Comercial",
    )

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "Title", parent=styles["Title"], fontName="Helvetica-Bold",
        fontSize=18, textColor=colors.HexColor("#1F4E78"), alignment=1,
        spaceAfter=10,
    )
    h2_style = ParagraphStyle(
        "H2", parent=styles["Heading2"], fontName="Helvetica-Bold",
        fontSize=12, textColor=colors.white, backColor=colors.HexColor("#2E75B6"),
        leftIndent=4, rightIndent=4, spaceAfter=4, spaceBefore=10,
        borderPadding=(4, 4, 4, 4),
    )
    desc_style = ParagraphStyle(
        "Desc", parent=styles["Normal"], fontName="Helvetica-Oblique",
        fontSize=9, textColor=colors.HexColor("#555555"), spaceAfter=4,
    )
    cell_style = ParagraphStyle(
        "Cell", parent=styles["Normal"], fontName="Helvetica",
        fontSize=8.5, leading=11,
    )
    cell_center = ParagraphStyle(
        "CellCenter", parent=cell_style, alignment=1,
    )
    info_style = ParagraphStyle(
        "Info", parent=styles["Normal"], fontName="Helvetica",
        fontSize=10, leading=14,
    )

    story = []

    # Portada
    story.append(Paragraph("Checklist de Conformidad — Módulo Comercial", title_style))
    story.append(Spacer(1, 6 * mm))

    info_data = [
        ["Nombre del responsable", ""],
        ["Rol", ""],
        ["Fecha de prueba", ""],
        ["Ambiente (Dev / QA / Prod)", ""],
    ]
    info_tbl = Table(info_data, colWidths=[60 * mm, 130 * mm])
    info_tbl.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (0, -1), colors.HexColor("#D9E1F2")),
        ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 10),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("LEFTPADDING", (0, 0), (-1, -1), 6),
        ("RIGHTPADDING", (0, 0), (-1, -1), 6),
        ("TOPPADDING", (0, 0), (-1, -1), 6),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
        ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#999999")),
    ]))
    story.append(info_tbl)
    story.append(Spacer(1, 6 * mm))

    legend_data = [
        ["Leyenda de estado", ""],
        ["[ ]", "Pendiente"],
        ["[x]", "Conforme"],
        ["[!]", "Con observación (anotar al final)"],
        ["[F]", "Falla bloqueante"],
    ]
    legend_tbl = Table(legend_data, colWidths=[30 * mm, 160 * mm])
    legend_tbl.setStyle(TableStyle([
        ("SPAN", (0, 0), (1, 0)),
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#305496")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("ALIGN", (0, 0), (-1, 0), "CENTER"),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("LEFTPADDING", (0, 0), (-1, -1), 6),
        ("RIGHTPADDING", (0, 0), (-1, -1), 6),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#999999")),
    ]))
    story.append(legend_tbl)
    story.append(PageBreak())

    # Secciones de casos
    page_width = landscape(A4)[0] - 24 * mm
    col_widths = [
        page_width * 0.06,  # #
        page_width * 0.27,  # Caso
        page_width * 0.34,  # Resultado esperado
        page_width * 0.10,  # Estado
        page_width * 0.23,  # Observación
    ]

    for sec in SECCIONES:
        story.append(Paragraph(sec["titulo"], h2_style))
        story.append(Paragraph(sec["descripcion"], desc_style))

        data = [["#", "Caso", "Resultado esperado", "Estado", "Observación"]]
        for num, caso, esperado in sec["casos"]:
            data.append([
                Paragraph(num, cell_center),
                Paragraph(caso, cell_style),
                Paragraph(esperado, cell_style),
                Paragraph("[ ]", cell_center),
                Paragraph("", cell_style),
            ])

        tbl = Table(data, colWidths=col_widths, repeatRows=1)
        tbl.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#305496")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, 0), 9),
            ("ALIGN", (0, 0), (-1, 0), "CENTER"),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("LEFTPADDING", (0, 0), (-1, -1), 4),
            ("RIGHTPADDING", (0, 0), (-1, -1), 4),
            ("TOPPADDING", (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#999999")),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F5F8FC")]),
        ]))
        story.append(tbl)
        story.append(Spacer(1, 4 * mm))

    # Resumen
    story.append(PageBreak())
    story.append(Paragraph("Resumen de conformidad", h2_style))

    sum_data = [["Sección", "Total casos", "Conformes", "Con obs.", "Bloqueantes"]]
    total_casos = 0
    for sec in SECCIONES:
        n = len(sec["casos"])
        total_casos += n
        sum_data.append([
            Paragraph(sec["titulo"], cell_style),
            Paragraph(str(n), cell_center),
            Paragraph("", cell_center),
            Paragraph("", cell_center),
            Paragraph("", cell_center),
        ])
    sum_data.append([
        Paragraph("<b>TOTAL</b>", cell_style),
        Paragraph(f"<b>{total_casos}</b>", cell_center),
        Paragraph("", cell_center),
        Paragraph("", cell_center),
        Paragraph("", cell_center),
    ])

    sum_widths = [page_width * 0.50, page_width * 0.12, page_width * 0.12, page_width * 0.12, page_width * 0.14]
    sum_tbl = Table(sum_data, colWidths=sum_widths, repeatRows=1)
    sum_tbl.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#305496")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 9),
        ("ALIGN", (0, 0), (-1, 0), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("BACKGROUND", (0, -1), (-1, -1), colors.HexColor("#D9E1F2")),
        ("LEFTPADDING", (0, 0), (-1, -1), 4),
        ("RIGHTPADDING", (0, 0), (-1, -1), 4),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#999999")),
    ]))
    story.append(sum_tbl)
    story.append(Spacer(1, 8 * mm))

    # Aprobación final
    story.append(Paragraph("Conformidad final", h2_style))
    aprob = [
        "[ ] <b>Aprobado</b> — todas las funcionalidades operan según lo esperado.",
        "[ ] <b>Aprobado con observaciones</b> — listadas en la sección anterior, no bloquean producción.",
        "[ ] <b>Rechazado</b> — existen fallas bloqueantes que impiden el pase a producción.",
    ]
    for txt in aprob:
        story.append(Paragraph(txt, info_style))
        story.append(Spacer(1, 2 * mm))

    story.append(Spacer(1, 12 * mm))
    story.append(Paragraph("Firma del responsable: __________________________________", info_style))
    story.append(Spacer(1, 6 * mm))
    story.append(Paragraph("Fecha: __________________________________", info_style))

    doc.build(story)
    print(f"PDF generado: {PDF_PATH}")


if __name__ == "__main__":
    generar_excel()
    generar_pdf()
